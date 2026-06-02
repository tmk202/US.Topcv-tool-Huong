#!/usr/bin/env python3
import argparse
import json
import os
import re
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from urllib import error, request

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STATE_FILE = ROOT / "excel-state.json"

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
PHONE_RE = re.compile(r"(?:\+?84|0)(?:[\s.-]?\d){8,10}")


@dataclass
class Candidate:
    job_id: str
    job_title: str
    apply_at: str
    candidate_name: str
    candidate_email: str
    candidate_phone: str
    download_url: str

    def key(self) -> str:
        parts = [
            self.job_id,
            self.candidate_email,
            digits_only(self.candidate_phone),
            self.candidate_name,
            self.apply_at,
        ]
        return "|".join(part.strip().lower() for part in parts if part and part.strip())

    def payload(self) -> dict[str, str]:
        return {
            "job_id": self.job_id,
            "job_title": self.job_title,
            "apply_at": self.apply_at,
            "candidate_name": self.candidate_name,
            "candidate_email": self.candidate_email,
            "candidate_phone": self.candidate_phone,
            "download_url": self.download_url,
        }


def load_env(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    replacements = {
        "à": "a", "á": "a", "ạ": "a", "ả": "a", "ã": "a",
        "â": "a", "ầ": "a", "ấ": "a", "ậ": "a", "ẩ": "a", "ẫ": "a",
        "ă": "a", "ằ": "a", "ắ": "a", "ặ": "a", "ẳ": "a", "ẵ": "a",
        "è": "e", "é": "e", "ẹ": "e", "ẻ": "e", "ẽ": "e",
        "ê": "e", "ề": "e", "ế": "e", "ệ": "e", "ể": "e", "ễ": "e",
        "ì": "i", "í": "i", "ị": "i", "ỉ": "i", "ĩ": "i",
        "ò": "o", "ó": "o", "ọ": "o", "ỏ": "o", "õ": "o",
        "ô": "o", "ồ": "o", "ố": "o", "ộ": "o", "ổ": "o", "ỗ": "o",
        "ơ": "o", "ờ": "o", "ớ": "o", "ợ": "o", "ở": "o", "ỡ": "o",
        "ù": "u", "ú": "u", "ụ": "u", "ủ": "u", "ũ": "u",
        "ư": "u", "ừ": "u", "ứ": "u", "ự": "u", "ử": "u", "ữ": "u",
        "ỳ": "y", "ý": "y", "ỵ": "y", "ỷ": "y", "ỹ": "y",
        "đ": "d",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def digits_only(value: str) -> str:
    return re.sub(r"\D+", "", value or "")


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def normalize_date(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    formats = [
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(value[:19], fmt)
            if fmt in {"%d/%m/%Y", "%Y-%m-%d"}:
                return dt.strftime("%Y-%m-%d 00:00:00")
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return value


def first_value(row: dict[str, str], patterns: tuple[str, ...]) -> str:
    for header, value in row.items():
        if value and any(pattern in header for pattern in patterns):
            return value
    return ""


def candidate_from_row(row: dict[str, str]) -> Optional[Candidate]:
    all_text = " ".join(row.values())
    email = first_value(row, ("email", "e mail", "thu dien tu"))
    phone = first_value(row, ("dien thoai", "so dien thoai", "phone", "mobile", "sdt"))
    if not email:
        match = EMAIL_RE.search(all_text)
        email = match.group(0) if match else ""
    if not phone:
        match = PHONE_RE.search(all_text)
        phone = match.group(0) if match else ""

    name = first_value(row, ("ho va ten", "ho ten", "ung vien", "ten ung vien", "candidate name", "full name"))
    job_title = first_value(row, ("chien dich", "vi tri", "tin tuyen dung", "cong viec", "job title", "campaign", "ten cdtd"))
    job_id = first_value(row, ("ten cdtd", "ma tin", "ma chien dich", "ma cdtd", "job id", "campaign id", "id tin"))
    apply_at = first_value(row, ("ngay ung tuyen", "ngay tiep nhan", "thoi gian ung tuyen", "apply", "applied", "created"))
    download_url = first_value(row, ("download", "tai cv", "link cv", "link xem cv", "url cv", "cv url", "resume"))

    if not download_url:
        for value in row.values():
            if "onetime-download" in value or "cv-management" in value:
                download_url = value
                break

    if not name and not email and not phone:
        return None

    return Candidate(
        job_id=job_id,
        job_title=job_title,
        apply_at=normalize_date(apply_at),
        candidate_name=name,
        candidate_email=email,
        candidate_phone=phone,
        download_url=download_url,
    )


def unique_header(header: str, seen: dict[str, int]) -> str:
    header = header or "column"
    count = seen.get(header, 0) + 1
    seen[header] = count
    if count == 1:
        return header
    return f"{header} {count}"


def read_topcv_records(xlsx_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = find_header_row(sheet)
    seen_headers: dict[str, int] = {}
    raw_headers = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
        (),
    )
    headers = []
    for idx, header in enumerate(raw_headers, start=1):
        headers.append(unique_header(cell_to_text(header) or f"Column {idx}", seen_headers))

    records: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = [cell_to_text(value) for value in row[: len(headers)]]
        if not any(values):
            continue
        records.append(dict(zip(headers, values)))
    return records


def find_header_row(sheet) -> int:
    best_row = 1
    best_score = -1
    for row_idx in range(1, min(sheet.max_row, 20) + 1):
        values = [normalize_header(sheet.cell(row_idx, col_idx).value) for col_idx in range(1, sheet.max_column + 1)]
        joined = " ".join(values)
        score = sum(
            token in joined
            for token in ("email", "dien thoai", "ung vien", "ho ten", "chien dich", "vi tri", "apply")
        )
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def read_candidates(xlsx_path: Path) -> list[Candidate]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = find_header_row(sheet)
    headers = [
        normalize_header(sheet.cell(header_row, col_idx).value) or f"column {col_idx}"
        for col_idx in range(1, sheet.max_column + 1)
    ]

    candidates: list[Candidate] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = [cell_to_text(value) for value in row[: len(headers)]]
        if not any(values):
            continue
        row = dict(zip(headers, values))
        candidate = candidate_from_row(row)
        if candidate:
            candidates.append(candidate)
    return candidates


def print_headers(xlsx_path: Path) -> None:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = find_header_row(sheet)
    headers = [
        cell_to_text(sheet.cell(header_row, col_idx).value)
        for col_idx in range(1, sheet.max_column + 1)
    ]
    print(f"Header row: {header_row}")
    for idx, header in enumerate(headers, start=1):
        print(f"{idx}: {header}")


def download_export(export_url: str) -> Path:
    temp = tempfile.NamedTemporaryFile(prefix="topcv-export-", suffix=".xlsx", delete=False)
    temp_path = Path(temp.name)
    temp.close()
    req = request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    with request.urlopen(req, timeout=120) as resp:
        content_type = resp.headers.get("content-type", "")
        data = resp.read()
        if resp.status >= 300:
            raise RuntimeError(f"TopCV export returned HTTP {resp.status}")
        if b"<html" in data[:500].lower():
            raise RuntimeError("TopCV export did not return an Excel file. The token may be expired.")
        temp_path.write_bytes(data)
        print(f"Downloaded export to {temp_path} ({content_type})")
    return temp_path


def load_state(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return set()
    if isinstance(data, list):
        return set(str(item) for item in data)
    return set(str(item) for item in data.get("sent", []))


def save_state(path: Path, sent: set[str]) -> None:
    path.write_text(json.dumps({"sent": sorted(sent)}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def post_lark(webhook_url: str, payload: dict[str, str]) -> None:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    retry_codes = {405, 408, 409, 425, 429, 500, 502, 503, 504}
    last_error = None

    for attempt in range(1, 5):
        req = request.Request(webhook_url, data=data, headers={"Content-Type": "application/json"}, method="POST")
        try:
            with request.urlopen(req, timeout=30) as resp:
                body = resp.read().decode("utf-8", errors="replace")
                if resp.status >= 300:
                    raise RuntimeError(f"Lark returned HTTP {resp.status}: {body[:1000]}")
                return
        except error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            headers = dict(exc.headers.items()) if exc.headers else {}
            detail = body[:1000] if body else "<empty body>"
            last_error = RuntimeError(
                f"Lark returned HTTP {exc.code} {exc.reason}: {detail}; headers={json.dumps(headers, ensure_ascii=False)}"
            )
            if exc.code not in retry_codes or attempt == 4:
                raise last_error from exc
            wait_seconds = attempt * 5
            print(f"Lark HTTP {exc.code}; retrying in {wait_seconds}s ({attempt}/4)...")
            time.sleep(wait_seconds)

    if last_error:
        raise last_error


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Send TopCV Excel export rows to Lark webhook.")
    parser.add_argument("--xlsx", help="Path to a downloaded TopCV .xlsx export file.")
    parser.add_argument("--export-url", help="TopCV export-excel URL. If omitted, TOPCV_EXPORT_URL is used.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print payloads without sending to Lark.")
    parser.add_argument("--print-headers", action="store_true", help="Print detected Excel headers and exit.")
    parser.add_argument("--reset-state", action="store_true", help="Ignore existing sent-state for this run.")
    return parser.parse_args()


def main() -> int:
    load_env(ROOT / ".env")
    args = parse_args()

    webhook_url = os.getenv("LARK_WEBHOOK_URL", "").strip()
    export_url = (args.export_url or os.getenv("TOPCV_EXPORT_URL", "")).strip()
    state_file = Path(os.getenv("EXCEL_STATE_FILE", str(DEFAULT_STATE_FILE))).expanduser()
    if not state_file.is_absolute():
        state_file = ROOT / state_file

    if not args.dry_run and not webhook_url:
        print("Missing LARK_WEBHOOK_URL. Add it to .env first.", file=sys.stderr)
        return 2

    if args.xlsx:
        xlsx_path = Path(args.xlsx).expanduser()
    elif export_url:
        xlsx_path = download_export(export_url)
    else:
        print("Provide --xlsx or set TOPCV_EXPORT_URL in .env.", file=sys.stderr)
        return 2

    if not xlsx_path.exists():
        print(f"Excel file not found: {xlsx_path}", file=sys.stderr)
        return 2

    if args.print_headers:
        print_headers(xlsx_path)
        return 0

    candidates = read_candidates(xlsx_path)
    sent = set() if args.reset_state else load_state(state_file)
    new_count = 0
    skipped_count = 0

    for candidate in candidates:
        key = candidate.key()
        if not key or key in sent:
            skipped_count += 1
            continue

        payload = candidate.payload()
        if args.dry_run:
            print(json.dumps(payload, ensure_ascii=False))
        else:
            post_lark(webhook_url, payload)
            print(f"Sent: {candidate.candidate_name} / {candidate.candidate_email}")

        sent.add(key)
        new_count += 1

    if not args.dry_run:
        save_state(state_file, sent)
    print(f"Done. Parsed {len(candidates)}, sent {new_count}, skipped {skipped_count}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
